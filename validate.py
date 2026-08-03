#!/usr/bin/env python3
"""Kiểm tra workbook Excel nghiệp vụ ngân hàng (4 sheet hồ sơ + sheet nghiệp vụ).

Chạy: python3 validate.py <file.xlsx> [--json]
Exit 0 = PASS, 1 = FAIL. Chỉ stdlib + openpyxl.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

PROFILE = frozenset({"META", "SOURCES", "DECISIONS", "CHECKS"})
FORMULA_ERRS = frozenset({"#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NAME?"})
# Id chỉ nhận chữ/số/gạch — KHÔNG dùng \S+ vì nó nuốt luôn dấu câu:
# comment thật "Căn cứ src:SRC01, đã đối chiếu" sẽ ra id "SRC01," rồi báo sai
# "không có trong SOURCES" (Codex review 2026-08-03, P1).
SRC_RE = re.compile(r"\bsrc:([A-Za-z0-9_-]+)", re.I)
DEC_RE = re.compile(r"\bdec:([A-Za-z0-9_-]+)", re.I)
# Ref ô: chấp cả 'Tên sheet có dấu'!B5 (nháy đôi '' = nháy trong tên) và $B$5.
CELL_RE = re.compile(
    r"^(?:'((?:[^']|'')+)'|([^'!]+))!\$?([A-Za-z]{1,3})\$?(\d+)$"
)


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _is_num(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _is_yellow(cell: Any) -> bool:
    fill = cell.fill
    if fill is None or fill.fgColor is None:
        return False
    rgb = getattr(fill.fgColor, "rgb", None)
    return bool(rgb) and str(rgb).upper().endswith("FFFF00")


def _comment(cell: Any) -> str:
    c = cell.comment
    return c.text if c and c.text else ""


def read_meta(wb: Workbook) -> dict[str, str]:
    out: dict[str, str] = {}
    if "META" not in wb.sheetnames:
        return out
    for a, b in wb["META"].iter_rows(min_row=1, max_col=2, values_only=True):
        if a is not None:
            out[_s(a)] = _s(b)
    return out


def source_ids(wb: Workbook) -> set[str]:
    if "SOURCES" not in wb.sheetnames:
        return set()
    return {
        _s(r[0])
        for r in wb["SOURCES"].iter_rows(min_row=2, max_col=1, values_only=True)
        if r[0]
    }


def decision_ids(wb: Workbook) -> set[str]:
    if "DECISIONS" not in wb.sheetnames:
        return set()
    return {
        _s(r[0])
        for r in wb["DECISIONS"].iter_rows(min_row=2, max_col=1, values_only=True)
        if r[0]
    }


def biz_sheets(wb: Workbook) -> list[Worksheet]:
    return [wb[n] for n in wb.sheetnames if n not in PROFILE]


def _ref(ws: Worksheet, row: int, col: int) -> str:
    return f"{ws.title}!{ws.cell(row, col).coordinate}"


def check_open_questions(wb: Workbook, meta: dict[str, str]) -> list[str]:
    errs: list[str] = []
    if "DECISIONS" not in wb.sheetnames:
        return errs
    ship = meta.get("trang_thai", "").lower() == "ship"
    for row in wb["DECISIONS"].iter_rows(min_row=2, max_col=8, values_only=True):
        did, status = _s(row[0]), _s(row[7]).lower()
        if not did or status != "open":
            continue
        if ship:
            errs.append(
                f"CRITICAL: DECISIONS còn open {did} trong khi META.trang_thai=ship"
            )
        else:
            errs.append(f"DECISIONS còn câu hỏi mở: {did}")
    return errs


def check_hardcode(wb: Workbook) -> list[str]:
    errs: list[str] = []
    for ws in biz_sheets(wb):
        for row in ws.iter_rows():
            for cell in row:
                if not _is_num(cell.value):
                    continue
                if _is_yellow(cell):
                    continue
                errs.append(
                    f"Nghi hardcode tại {_ref(ws, cell.row, cell.column)} = {cell.value}"
                )
    return errs


def check_provenance(wb: Workbook) -> list[str]:
    errs: list[str] = []
    srcs, decs = source_ids(wb), decision_ids(wb)
    for ws in biz_sheets(wb):
        for row in ws.iter_rows():
            for cell in row:
                if not _is_yellow(cell):
                    continue
                if cell.value is None or cell.value == "":
                    continue
                text = _comment(cell)
                sm, dm = SRC_RE.search(text), DEC_RE.search(text)
                ref = _ref(ws, cell.row, cell.column)
                if not sm and not dm:
                    errs.append(f"Ô vàng {ref} thiếu comment src:<id> hoặc dec:<id>")
                    continue
                if sm and sm.group(1) not in srcs:
                    errs.append(
                        f"Ô vàng {ref}: source_id '{sm.group(1)}' không có trong SOURCES"
                    )
                if dm and dm.group(1) not in decs:
                    errs.append(
                        f"Ô vàng {ref}: decision_id '{dm.group(1)}' không có trong DECISIONS"
                    )
    return errs


def check_money_integrity(wb_val: Workbook) -> list[str]:
    errs: list[str] = []
    for ws in biz_sheets(wb_val):
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not _is_num(v):
                    continue
                if abs(v) > 1000 and abs(v - round(v)) > 1e-9:
                    errs.append(
                        f"Tiền chưa làm tròn tại {_ref(ws, cell.row, cell.column)} = {v}"
                    )
    return errs


def check_unit_label(wb_val: Workbook) -> list[str]:
    warns: list[str] = []
    for ws in biz_sheets(wb_val):
        has_big = any(
            _is_num(c.value) and abs(c.value) > 1_000_000
            for row in ws.iter_rows()
            for c in row
        )
        if not has_big:
            continue
        head = " ".join(
            _s(c.value)
            for r in ws.iter_rows(min_row=1, max_row=10)
            for c in r
            if c.value is not None
        )
        if "Đơn vị" not in head:
            warns.append(
                f"Sheet '{ws.title}' có số > 10^6 nhưng 10 dòng đầu thiếu nhãn 'Đơn vị'"
            )
    return warns


def _cell_val(wb: Workbook, ref: str) -> Any:
    m = CELL_RE.match(ref.strip())
    if not m:
        return None
    quoted, plain, col, row = m.group(1), m.group(2), m.group(3), m.group(4)
    # Tên sheet trong nháy: Excel escape nháy đơn bằng cách nhân đôi ('O''Brien').
    sheet = quoted.replace("''", "'") if quoted is not None else plain.strip()
    if sheet not in wb.sheetnames:
        return None
    return wb[sheet][f"{col.upper()}{row}"].value


def check_balance(wb_val: Workbook, meta: dict[str, str]) -> list[str]:
    errs: list[str] = []
    raw = meta.get("balance_pairs", "")
    if not raw:
        return errs
    for part in raw.split(";"):
        part = part.strip()
        if not part or "=" not in part:
            continue
        left, right = [x.strip() for x in part.split("=", 1)]
        lv, rv = _cell_val(wb_val, left), _cell_val(wb_val, right)
        if not _is_num(lv) or not _is_num(rv):
            errs.append(
                f"Cặp cân đối {left}={right}: không đọc được số "
                f"(left={lv!r}, right={rv!r})"
            )
            continue
        diff = abs(float(lv) - float(rv))
        if diff > 1:
            errs.append(
                f"Lệch cân đối {left}={right}: lệch {diff:g} "
                f"(trái={lv:g}, phải={rv:g})"
            )
    return errs


def check_formula_errors(wb_val: Workbook) -> list[str]:
    errs: list[str] = []
    for name in wb_val.sheetnames:
        ws = wb_val[name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.strip() in FORMULA_ERRS:
                    errs.append(
                        f"Lỗi công thức {v.strip()} tại {name}!{cell.coordinate}"
                    )
    return errs


def run_checks(path: Path) -> tuple[list[dict[str, str]], int]:
    wb = load_workbook(path, data_only=False)
    wb_val = load_workbook(path, data_only=True)
    meta = read_meta(wb)
    rules: list[tuple[str, list[str], str]] = [
        ("check_open_questions", check_open_questions(wb, meta), "FAIL"),
        ("check_hardcode", check_hardcode(wb), "FAIL"),
        ("check_provenance", check_provenance(wb), "FAIL"),
        ("check_money_integrity", check_money_integrity(wb_val), "FAIL"),
        ("check_unit_label", check_unit_label(wb_val), "WARN"),
        ("check_balance", check_balance(wb_val, meta), "FAIL"),
        ("check_formula_errors", check_formula_errors(wb_val), "FAIL"),
    ]
    rows: list[dict[str, str]] = []
    n_fail = 0
    for name, msgs, severity in rules:
        if not msgs:
            rows.append({"rule": name, "status": "PASS", "detail": ""})
        else:
            if severity == "FAIL":
                n_fail += 1
            rows.append({"rule": name, "status": severity, "detail": "; ".join(msgs)})
    return rows, n_fail


def append_checks(path: Path, passed: bool, n_errors: int) -> None:
    wb = load_workbook(path)
    if "CHECKS" not in wb.sheetnames:
        ws = wb.create_sheet("CHECKS")
        ws.append(["timestamp", "script", "ket_qua", "so_loi"])
    else:
        ws = wb["CHECKS"]
        if ws.max_row == 0 or ws.cell(1, 1).value is None:
            ws.append(["timestamp", "script", "ket_qua", "so_loi"])
    ws.append([
        datetime.now().isoformat(timespec="seconds"),
        "validate.py",
        "PASS" if passed else "FAIL",
        n_errors,
    ])
    wb.save(path)


def format_table(rows: list[dict[str, str]]) -> str:
    lines = [
        f"{'Luật':<28} | {'Trạng thái':<10} | Chi tiết",
        f"{'-' * 28}-+-{'-' * 10}-+-{'-' * 40}",
    ]
    for r in rows:
        lines.append(f"{r['rule']:<28} | {r['status']:<10} | {r['detail']}")
    return "\n".join(lines)


def validate(path: Path) -> tuple[bool, list[dict[str, str]], int]:
    rows, n_fail = run_checks(path)
    passed = n_fail == 0
    n_err_msgs = 0
    for r in rows:
        if r["status"] == "FAIL" and r["detail"]:
            n_err_msgs += r["detail"].count("; ") + 1
    append_checks(path, passed, n_err_msgs if not passed else 0)
    return passed, rows, n_err_msgs


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Validate workbook Excel nghiệp vụ ngân hàng")
    p.add_argument("file", type=Path, help="Đường dẫn file .xlsx")
    p.add_argument("--json", action="store_true", help="In kết quả JSON")
    args = p.parse_args(argv)
    if not args.file.is_file():
        print(f"Không tìm thấy file: {args.file}", file=sys.stderr)
        return 1
    passed, rows, n_err = validate(args.file)
    status = "PASS" if passed else "FAIL"
    if args.json:
        print(json.dumps(
            {"status": status, "n_errors": n_err, "rules": rows},
            ensure_ascii=False, indent=2,
        ))
    else:
        print(f"Kết quả: {status}")
        print(format_table(rows))
    return 0 if passed else 1


if __name__ == "__main__":
    raise SystemExit(main())
