"""Pytest cho validate.py — tự sinh fixture xlsx bằng openpyxl."""
from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import validate  # noqa: E402

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def _profile(wb: Workbook, *, trang_thai: str = "draft", balance_pairs: str = ""):
    meta = wb.create_sheet("META", 0)
    for k, v in [
        ("method_version", "1.0"),
        ("method_hash", "abc"),
        ("ky", "2026-07"),
        ("cutoff", "2026-07-31"),
        ("don_vi", "VNĐ"),
        ("nguoi_ky", "Tester"),
        ("trang_thai", trang_thai),
    ]:
        meta.append([k, v])
    if balance_pairs:
        meta.append(["balance_pairs", balance_pairs])

    src = wb.create_sheet("SOURCES")
    src.append(["source_id", "mo_ta", "file", "ngay_nhan", "ghi_chu"])
    src.append(["SRC01", "Bảng cân đối", "bcd.xlsx", "2026-07-01", ""])

    dec = wb.create_sheet("DECISIONS")
    dec.append([
        "decision_id", "cau_hoi", "dap_an", "scope",
        "het_han", "nguoi_ky", "ngay", "trang_thai",
    ])
    return dec


def _checks(wb: Workbook) -> None:
    ch = wb.create_sheet("CHECKS")
    ch.append(["timestamp", "script", "ket_qua", "so_loi"])


def make_pass_wb(path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    dec = _profile(wb, trang_thai="draft", balance_pairs="BC!B5=BC!B6")
    dec.append(["DEC01", "Có điều chỉnh?", "Không", "ky-nay", "", "A", "2026-07-01", "resolved"])
    _checks(wb)
    bc = wb.create_sheet("BC")
    bc["A1"] = "Đơn vị: VNĐ"
    bc["B2"] = 500_000
    bc["B2"].fill = YELLOW
    bc["B2"].comment = Comment("src:SRC01", "test")
    bc["B3"] = "=B2*2"
    bc["B5"] = 100
    bc["B5"].fill = YELLOW
    bc["B5"].comment = Comment("src:SRC01", "test")
    bc["B6"] = 100
    bc["B6"].fill = YELLOW
    bc["B6"].comment = Comment("dec:DEC01", "test")
    bc["C2"] = 12.5
    bc["C2"].fill = YELLOW
    bc["C2"].comment = Comment("src:SRC01", "test")
    wb.save(path)
    return path


def make_fail_hardcode_open(path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    dec = _profile(wb, trang_thai="draft")
    dec.append(["DEC02", "Chưa trả lời?", "", "ky-nay", "", "A", "2026-07-01", "open"])
    _checks(wb)
    bc = wb.create_sheet("BC")
    bc["A1"] = "Đơn vị: VNĐ"
    bc["B2"] = 999  # hardcode
    bc["B3"] = 10
    bc["B3"].fill = YELLOW
    bc["B3"].comment = Comment("src:SRC01", "test")
    wb.save(path)
    return path


def make_fail_balance_money(path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    dec = _profile(wb, trang_thai="draft", balance_pairs="BC!B5=BC!B6")
    dec.append(["DEC01", "OK?", "Có", "vinh-vien", "", "A", "2026-07-01", "resolved"])
    _checks(wb)
    bc = wb.create_sheet("BC")
    bc["A1"] = "Đơn vị: VNĐ"
    bc["B5"] = 1000
    bc["B5"].fill = YELLOW
    bc["B5"].comment = Comment("src:SRC01", "test")
    bc["B6"] = 1050
    bc["B6"].fill = YELLOW
    bc["B6"].comment = Comment("src:SRC01", "test")
    bc["C2"] = 1500.75
    bc["C2"].fill = YELLOW
    bc["C2"].comment = Comment("src:SRC01", "test")
    wb.save(path)
    return path


@pytest.fixture
def pass_wb(tmp_path: Path) -> Path:
    return make_pass_wb(tmp_path / "pass.xlsx")


@pytest.fixture
def fail_hardcode_open_wb(tmp_path: Path) -> Path:
    return make_fail_hardcode_open(tmp_path / "fail_ho.xlsx")


@pytest.fixture
def fail_balance_money_wb(tmp_path: Path) -> Path:
    return make_fail_balance_money(tmp_path / "fail_bm.xlsx")


def test_pass_exit_0(pass_wb: Path) -> None:
    assert validate.main([str(pass_wb)]) == 0


def test_fail_hardcode_and_open_question(fail_hardcode_open_wb: Path) -> None:
    passed, rows, _ = validate.validate(fail_hardcode_open_wb)
    assert passed is False
    by = {r["rule"]: r for r in rows}
    assert by["check_open_questions"]["status"] == "FAIL"
    assert by["check_hardcode"]["status"] == "FAIL"
    assert "DEC02" in by["check_open_questions"]["detail"]
    assert "BC!B2" in by["check_hardcode"]["detail"]
    assert validate.main([str(fail_hardcode_open_wb)]) == 1


def test_fail_balance_and_money(fail_balance_money_wb: Path) -> None:
    passed, rows, _ = validate.validate(fail_balance_money_wb)
    assert passed is False
    by = {r["rule"]: r for r in rows}
    assert by["check_balance"]["status"] == "FAIL"
    assert by["check_money_integrity"]["status"] == "FAIL"
    assert "BC!B5" in by["check_balance"]["detail"] or "BC!B6" in by["check_balance"]["detail"]
    assert "BC!C2" in by["check_money_integrity"]["detail"]
    assert validate.main([str(fail_balance_money_wb)]) == 1


def test_comment_co_dau_cau_van_lay_dung_id() -> None:
    # Codex review 03/08 (P1): \S+ nuốt cả dấu câu → id "SRC01," lệch set SOURCES.
    for text, want in [
        ("src:SRC01", "SRC01"),
        ("Căn cứ src:SRC01, đã đối chiếu", "SRC01"),
        ("(src:SRC01)", "SRC01"),
        ("src:SRC01; dec:DEC01", "SRC01"),
        ("Theo src:SRC01. Hết", "SRC01"),
    ]:
        m = validate.SRC_RE.search(text)
        assert m and m.group(1) == want, f"{text!r} → {m and m.group(1)!r}"
    m = validate.DEC_RE.search("src:SRC01; dec:DEC01.")
    assert m and m.group(1) == "DEC01"


def test_cell_ref_tuyet_doi_va_ten_sheet_co_nhay(tmp_path: Path) -> None:
    # Codex review 03/08 (P1): $B$5 và 'Tên sheet'!B5 trước đây trả None câm lặng.
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("BC")
    ws["B5"] = 100
    ws2 = wb.create_sheet("Báo cáo tháng 7")
    ws2["C3"] = 250
    p = tmp_path / "refs.xlsx"
    wb.save(p)

    from openpyxl import load_workbook
    loaded = load_workbook(p, data_only=True)
    assert validate._cell_val(loaded, "BC!B5") == 100
    assert validate._cell_val(loaded, "BC!$B$5") == 100
    assert validate._cell_val(loaded, "BC!b5") == 100
    assert validate._cell_val(loaded, "'Báo cáo tháng 7'!C3") == 250
    assert validate._cell_val(loaded, "Báo cáo tháng 7!C3") == 250
    assert validate._cell_val(loaded, "KHONGCO!B5") is None
    assert validate._cell_val(loaded, "rác") is None


def test_json_flag(pass_wb: Path, capsys: pytest.CaptureFixture[str]) -> None:
    code = validate.main([str(pass_wb), "--json"])
    assert code == 0
    data = json.loads(capsys.readouterr().out)
    assert data["status"] == "PASS"


def test_format_table_xuong_dong_that() -> None:
    # Cùng họ lỗi double-escape: "\\n".join in ra ký tự \n literal, bảng dính 1 dòng.
    txt = validate.format_table([
        {"rule": "check_hardcode", "status": "PASS", "detail": ""},
        {"rule": "check_balance", "status": "FAIL", "detail": "lệch 50"},
    ])
    assert "\\n" not in txt
    assert len(txt.splitlines()) == 4  # header + kẻ ngang + 2 dòng luật
